# app/routes/reportes.py

from datetime import date
from flask import Blueprint, render_template
from flask_login import login_required, current_user
from app.utils.roles  import rol_requerido
from app.utils.fechas import contar_habiles, dias_habiles_mes
from app.models.vendedor import Vendedor
from app.models.ventas   import BDVenta
from io import BytesIO
from datetime import datetime
from flask import Blueprint, request, send_file, flash, redirect, url_for
from sqlalchemy import case, literal
import pandas as pd
from openpyxl.utils.cell import get_column_letter
from decimal import Decimal

from app import db
from app.models.pedidos import BDPedido
from app.models.pedido_item import BDPedidoItem
from app.models.vendedor import Vendedor
from app.models.producto  import Producto

reportes_bp = Blueprint('reportes', __name__, url_prefix='/reportes')

@reportes_bp.route('/panel', methods=['GET'])
@login_required
@rol_requerido('semiadmin','administrador')
def panel():
    hoy = date.today()
    inicio_mes = hoy.replace(day=1)

    # Cálculo de días hábiles
    dias_transcurridos = contar_habiles(inicio_mes, hoy)
    total_habiles      = dias_habiles_mes(hoy.year, hoy.month)

    rows = []
    for vend in Vendedor.query.order_by(Vendedor.nombre).all():
        ventas = BDVenta.query.filter(
            BDVenta.codigo_vendedor==vend.codigo_vendedor,
            BDVenta.fecha >= inicio_mes,
            BDVenta.fecha <= hoy
        ).all()

        tot_vendido = sum(v.total_venta for v in ventas)
        tot_com     = sum(v.comision    for v in ventas)
        proy = (tot_vendido / dias_transcurridos * total_habiles
                if dias_transcurridos else 0)

        rows.append({
            'vendedor':   vend.nombre,
            'cod_vend':   vend.codigo_vendedor,
            'vendido':    tot_vendido,
            'comision':   tot_com,
            'proyeccion': proy
        })

    return render_template(
        'reportes/panel.html',
        rows=rows,
        dias_transcurridos=dias_transcurridos,
        total_habiles=total_habiles
    )

# 1) Formulario para pedidos por producto
@reportes_bp.route('/pedidos_por_producto', methods=['GET'])
@login_required
@rol_requerido('administrador','semiadmin')
def pedidos_por_producto():
    return render_template('reportes/pedidos_por_producto.html')

# 2) Exportación a Excel de pedidos por producto
@reportes_bp.route('/pedidos_por_producto/export', methods=['GET'])
@login_required
@rol_requerido('administrador','semiadmin')
def export_pedidos_productos_excel():
    # 1) Leer y validar fechas
    start = request.args.get('start', '').strip()
    end   = request.args.get('end',   '').strip()
    try:
        start_date = datetime.strptime(start, '%Y-%m-%d').date()
        end_date   = datetime.strptime(end,   '%Y-%m-%d').date()
    except ValueError:
        flash('Rango de fechas inválido.', 'warning')
        return redirect(url_for('reportes.pedidos_por_producto'))

    # 2) Recorrer pedidos en el rango y preparar filas
    rows = []
    pedidos = BDPedido.query.filter(
        BDPedido.fecha >= start_date,
        BDPedido.fecha <= end_date
    ).all()

    for ped in pedidos:
        vend = Vendedor.query.filter_by(codigo_vendedor=ped.codigo_vendedor).first()
        for item in ped.items:
            prod = Producto.query.filter_by(codigo=item.producto_cod).first()
            cat = (prod.categoria or '').lower()
            pct = ((vend.comision_panaderia if cat=='panadería' else vend.comision_bizcocheria) or 0)/100.0
            valor_total = float(item.subtotal)
            valor_neto  = valor_total * (1.0 - pct)

            rows.append({
                'Fecha':                        ped.fecha,
                'Año':                          ped.fecha.year,
                'cod vendedor':                 ped.codigo_vendedor,
                'nombre vendedor':              vend.nombre,
                'ruta':                         '',
                'codigo producto':              item.producto_cod,
                'nombre producto':              prod.nombre,
                'cantidad':                     item.cantidad,
                'valor total':                  valor_total,
                'valor neto (menos comisión)':  valor_neto,
                'lote':                         '',
                'mes':                          ped.fecha.month,
                'día':                          ped.fecha.day,
                'nombre del día':               ped.fecha.strftime('%A')
            })

    # 3) Crear DataFrame
    df = pd.DataFrame(rows)
    if df.empty:
        flash('No hay datos en ese rango.', 'info')
        return redirect(url_for('reportes.pedidos_por_producto'))
    df['Fecha'] = pd.to_datetime(df['Fecha'])

    # 4) Generar Excel en memoria usando openpyxl
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='PedidosPorProducto')
        ws = writer.sheets['PedidosPorProducto']
        for idx, col in enumerate(df.columns, start=1):
            max_len = max(df[col].astype(str).map(len).max(), len(col))
            ws.column_dimensions[get_column_letter(idx)].width = max_len + 2
    output.seek(0)

    # 5) Enviar al cliente
    filename = f"PedidosPorProducto_{start}_a_{end}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# 3) Formulario para pedidos día a día por vendedor
@reportes_bp.route('/pedidos_dia', methods=['GET'])
@login_required
@rol_requerido('administrador','semiadmin')
def pedidos_dia_por_vendedor():
    return render_template('reportes/pedidos_dia_por_vendedor.html')

# 4) Exportación a Excel de pedidos día a día
@reportes_bp.route('/pedidos_dia/export', methods=['GET'])
@login_required
@rol_requerido('administrador','semiadmin')
def export_pedidos_dia_excel():
    # 1) Leer y validar fechas
    start = request.args.get('start','').strip()
    end   = request.args.get('end','').strip()
    try:
        dt_start = datetime.strptime(start, '%Y-%m-%d').date()
        dt_end   = datetime.strptime(end,   '%Y-%m-%d').date()
    except ValueError:
        flash("Rango de fechas inválido.", "danger")
        return redirect(url_for('reportes.pedidos_dia_por_vendedor'))

    # 2) Consultar con ORM: sumar subtotal por fecha y vendedor
    q = (
        db.session.query(
            BDPedido.fecha.label('Fecha'),
            Vendedor.nombre.label('Vendedor'),
            db.func.sum(BDPedidoItem.subtotal).label('Total')
        )
        .join(BDPedidoItem, BDPedidoItem.pedido_id == BDPedido.id)
        .join(Vendedor, Vendedor.codigo_vendedor == BDPedido.codigo_vendedor)
        .filter(BDPedido.fecha >= dt_start, BDPedido.fecha <= dt_end)
        .group_by(BDPedido.fecha, Vendedor.nombre)
        .order_by(BDPedido.fecha, Vendedor.nombre)
    )

    rows = [
        {'Fecha': rec.Fecha, 'Vendedor': rec.Vendedor, 'Total': float(rec.Total)}
        for rec in q
    ]
    if not rows:
        flash("No hay datos para ese rango.", "info")
        return redirect(url_for('reportes.pedidos_dia_por_vendedor'))

    # 3) DataFrame y pivot: filas=Fecha, cols=Vendedor
    df = pd.DataFrame(rows)
    df['Fecha'] = pd.to_datetime(df['Fecha'])
    table = df.pivot(index='Fecha', columns='Vendedor', values='Total').fillna(0)

    # 4) Totales y proyección
    table.loc['Total'] = table.sum(axis=0)
    dias_hab  = contar_habiles(dt_start, dt_end)
    tot_hab   = dias_habiles_mes(dt_start.year, dt_start.month)
    if dias_hab:
        proj = table.loc['Total'] / dias_hab * tot_hab
    else:
        proj = table.loc['Total'] * 0
    table.loc['Proyección'] = proj

    # 5) Exportar a Excel con OpenPyXL
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        table.to_excel(writer, sheet_name='PedidosDia')
        ws = writer.sheets['PedidosDia']
        # Ajustar ancho de columna
        df_reset = table.reset_index()
        for idx, col in enumerate(df_reset.columns, start=1):
            max_len = max(df_reset[col].astype(str).map(len).max(), len(str(col)))
            ws.column_dimensions[get_column_letter(idx)].width = max_len + 2
    output.seek(0)

    # 6) Devolver archivo
    filename = f"PedidosDia_{start}_a_{end}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# 5) Formulario para pedidos mes a mes por vendedor
@reportes_bp.route('/pedidos_mes', methods=['GET'])
@login_required
@rol_requerido('administrador','semiadmin')
def pedidos_mes_por_vendedor():
    # pasamos la función date.today() para mostrar el año actual en el form
    return render_template('reportes/pedidos_mes_por_vendedor.html', now=date.today)

# 6) Exportación a Excel de pedidos mes a mes
@reportes_bp.route('/pedidos_mes/export', methods=['GET'])
@login_required
@rol_requerido('administrador','semiadmin')
def export_pedidos_mes_excel():
    # 1) Leer y validar año
    year_str = request.args.get('year','').strip()
    try:
        year = int(year_str)
    except ValueError:
        flash("Año inválido.", "danger")
        return redirect(url_for('reportes.pedidos_mes_por_vendedor'))

    # 2) Consulta con SQLAlchemy ORM
    q = (
        db.session.query(
            db.func.extract('month', BDPedido.fecha).label('Mes'),
            Vendedor.nombre.label('Vendedor'),
            db.func.sum(BDPedidoItem.subtotal).label('Total')
        )
        .join(BDPedidoItem, BDPedidoItem.pedido_id==BDPedido.id)
        .join(Vendedor, Vendedor.codigo_vendedor==BDPedido.codigo_vendedor)
        .filter(db.func.extract('year', BDPedido.fecha)==year)
        .group_by('Mes','Vendedor')
        .order_by('Mes','Vendedor')
    )

    rows = []
    for mes, vendedor, total in q:
        rows.append({
            'Mes': int(mes),
            'Vendedor': vendedor,
            'Total': float(total)
        })

    if not rows:
        flash(f"No hay datos para el año {year}.", "info")
        return redirect(url_for('reportes.pedidos_mes_por_vendedor'))

    # 3) DataFrame y pivot
    df = pd.DataFrame(rows)
    table = df.pivot(index='Mes', columns='Vendedor', values='Total').fillna(0)

    # 4) Totales y Proyección sin usar append()
    # Fila de totales
    table.loc['Total'] = table.sum(axis=0)
    # Fila de proyección anual
    meses_pasados = date.today().month if date.today().year == year else 12
    if meses_pasados:
        proj_values = table.loc['Total'] / meses_pasados * 12
    else:
        proj_values = table.loc['Total'] * 0
    # Asignamos directamente
    table.loc['Proyección'] = proj_values

    # 5) Exportar a Excel con openpyxl
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        table.to_excel(writer, sheet_name=f'Pedidos_{year}')
        ws = writer.sheets[f'Pedidos_{year}']
        df_reset = table.reset_index()
        for idx, col in enumerate(df_reset.columns, start=1):
            max_len = max(df_reset[col].astype(str).map(len).max(), len(str(col)))
            ws.column_dimensions[get_column_letter(idx)].width = max_len + 2
    output.seek(0)

    # 6) Enviar archivo
    filename = f"PedidosMes_{year}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@reportes_bp.route('/mi_panel', methods=['GET'])
@login_required
@rol_requerido('vendedor')
def mi_panel():
    from datetime import date
    from flask import request
    from app.models.ventas     import BDVenta
    from app.models.pedidos    import BDPedido
    from app.models.extras     import BDExtra  # ← Asegúrate que esté importado
    from app.models.cambio     import BD_CAMBIO
    from app.utils.fechas      import contar_habiles, dias_habiles_mes

    hoy = date.today()

    # Leer filtros GET con valores por defecto
    mes = int(request.args.get('mes', hoy.month))
    anio = int(request.args.get('anio', hoy.year))

    # Establecer inicio y fin del mes
    inicio_mes = date(anio, mes, 1)
    if mes == 12:
        fin_mes = date(anio + 1, 1, 1)
    else:
        fin_mes = date(anio, mes + 1, 1)

    # Días hábiles
    dias_transcurridos = contar_habiles(inicio_mes, hoy if (anio == hoy.year and mes == hoy.month) else fin_mes)
    total_habiles = dias_habiles_mes(anio, mes)

    # Ventas del mes
    ventas = BDVenta.query.filter(
        BDVenta.codigo_vendedor == current_user.codigo_vendedor,
        BDVenta.fecha >= inicio_mes,
        BDVenta.fecha < fin_mes
    ).all()

    tot_vendido = sum(v.total_venta for v in ventas)
    tot_com     = sum(v.comision    for v in ventas)
    proy = (tot_vendido / dias_transcurridos * total_habiles) if dias_transcurridos else 0

    # Total de pedidos del mes
    pedidos = BDPedido.query.filter_by(codigo_vendedor=current_user.codigo_vendedor).filter(
        BDPedido.fecha >= inicio_mes,
        BDPedido.fecha < fin_mes
    ).all()
    total_pedidos = sum(
        sum(item.subtotal for item in p.items) for p in pedidos
    )

    # Total de extras del mes
    extras = BDExtra.query.filter_by(codigo_vendedor=current_user.codigo_vendedor).filter(
        BDExtra.fecha >= inicio_mes,
        BDExtra.fecha < fin_mes
    ).all()
    total_extras = sum(
        sum(item.subtotal for item in e.items) for e in extras
    )

    # Sumar ambos totales
    total_pedidos += total_extras

    # Total de cambios (valor_cambio) del mes
    cambios = BD_CAMBIO.query.filter_by(codigo_vendedor=current_user.codigo_vendedor).filter(
        BD_CAMBIO.fecha >= inicio_mes,
        BD_CAMBIO.fecha < fin_mes
    ).all()
    total_cambios = sum(c.valor_cambio for c in cambios)

    # Renderizar plantilla
    return render_template(
        'reportes/mi_panel.html',
        anio=anio,
        mes=mes,
        vendido=tot_vendido,
        comision=tot_com,
        proyeccion=proy,
        total_pedidos=total_pedidos,
        total_cambios=total_cambios,
        dias_transcurridos=dias_transcurridos,
        total_habiles=total_habiles
    )

# Formulario de extras por producto
@reportes_bp.route('/extra_por_producto', methods=['GET'])
@login_required
@rol_requerido('administrador','semiadmin')
def extras_por_producto():
    return render_template('reportes/extra_por_producto.html')

# Exportación a Excel de extras por producto
@reportes_bp.route('/extra_por_producto/export', methods=['GET'])
@login_required
@rol_requerido('administrador','semiadmin')
def export_extras_productos_excel():
    from app.models.extras import BDExtra
    from app.models.extra_item import BDExtraItem
    from app.models.vendedor import Vendedor
    from app.models.producto import Producto

    start = request.args.get('start', '').strip()
    end   = request.args.get('end', '').strip()
    try:
        start_date = datetime.strptime(start, '%Y-%m-%d').date()
        end_date   = datetime.strptime(end, '%Y-%m-%d').date()
    except ValueError:
        flash('Rango de fechas inválido.', 'warning')
        return redirect(url_for('reportes.extras_por_producto'))

    rows = []
    extras = BDExtra.query.filter(
        BDExtra.fecha >= start_date,
        BDExtra.fecha <= end_date
    ).all()

    for ex in extras:
        vend = Vendedor.query.filter_by(codigo_vendedor=ex.codigo_vendedor).first()
        for item in ex.items:
            prod = Producto.query.filter_by(codigo=item.producto_cod).first()
            cat = (prod.categoria or '').lower()
            pct = ((vend.comision_panaderia if cat=='panadería' else vend.comision_bizcocheria) or 0)/100.0
            valor_total = float(item.subtotal)
            valor_neto  = valor_total * (1.0 - pct)

            rows.append({
                'Fecha':                        ex.fecha,
                'Año':                          ex.fecha.year,
                'cod vendedor':                 ex.codigo_vendedor,
                'nombre vendedor':              vend.nombre,
                'ruta':                         '',
                'codigo producto':              item.producto_cod,
                'nombre producto':              prod.nombre,
                'cantidad':                     item.cantidad,
                'valor total':                  valor_total,
                'valor neto (menos comisión)':  valor_neto,
                'lote':                         '',
                'mes':                          ex.fecha.month,
                'día':                          ex.fecha.day,
                'nombre del día':               ex.fecha.strftime('%A')
            })

    df = pd.DataFrame(rows)
    if df.empty:
        flash('No hay datos en ese rango.', 'info')
        return redirect(url_for('reportes.extras_por_producto'))
    df['Fecha'] = pd.to_datetime(df['Fecha'])

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='ExtrasPorProducto')
        ws = writer.sheets['ExtrasPorProducto']
        for idx, col in enumerate(df.columns, start=1):
            max_len = max(df[col].astype(str).map(len).max(), len(col))
            ws.column_dimensions[get_column_letter(idx)].width = max_len + 2
    output.seek(0)

    filename = f"ExtrasPorProducto_{start}_a_{end}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# Formulario de devoluciones por producto
@reportes_bp.route('/devoluciones_por_producto', methods=['GET'])
@login_required
@rol_requerido('administrador','semiadmin')
def devoluciones_por_producto():
    return render_template('reportes/devoluciones_por_producto.html')

# Exportación a Excel de devoluciones por producto
@reportes_bp.route('/devoluciones_por_producto/export', methods=['GET'])
@login_required
@rol_requerido('administrador','semiadmin')
def export_devoluciones_productos_excel():
    from app.models.devoluciones import BDDevolucion
    from app.models.devolucion_item import BDDevolucionItem
    from app.models.vendedor import Vendedor
    from app.models.producto import Producto

    start = request.args.get('start', '').strip()
    end   = request.args.get('end', '').strip()
    try:
        start_date = datetime.strptime(start, '%Y-%m-%d').date()
        end_date   = datetime.strptime(end, '%Y-%m-%d').date()
    except ValueError:
        flash('Rango de fechas inválido.', 'warning')
        return redirect(url_for('reportes.devoluciones_por_producto'))

    rows = []
    devols = BDDevolucion.query.filter(
        BDDevolucion.fecha >= start_date,
        BDDevolucion.fecha <= end_date
    ).all()

    for dev in devols:
        vend = Vendedor.query.filter_by(codigo_vendedor=dev.codigo_vendedor).first()
        for item in dev.items:
            prod = Producto.query.filter_by(codigo=item.producto_cod).first()
            cat = (prod.categoria or '').lower()
            pct = ((vend.comision_panaderia if cat=='panadería' else vend.comision_bizcocheria) or 0)/100.0
            valor_total = float(item.subtotal)
            valor_neto  = valor_total * (1.0 - pct)

            rows.append({
                'Fecha':                        dev.fecha,
                'Año':                          dev.fecha.year,
                'cod vendedor':                 dev.codigo_vendedor,
                'nombre vendedor':              vend.nombre,
                'ruta':                         '',
                'codigo producto':              item.producto_cod,
                'nombre producto':              prod.nombre,
                'cantidad':                     item.cantidad,
                'valor total':                  valor_total,
                'valor neto (menos comisión)':  valor_neto,
                'lote':                         '',
                'mes':                          dev.fecha.month,
                'día':                          dev.fecha.day,
                'nombre del día':               dev.fecha.strftime('%A')
            })

    df = pd.DataFrame(rows)
    if df.empty:
        flash('No hay datos en ese rango.', 'info')
        return redirect(url_for('reportes.devoluciones_por_producto'))
    df['Fecha'] = pd.to_datetime(df['Fecha'])

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='DevolucionesPorProducto')
        ws = writer.sheets['DevolucionesPorProducto']
        for idx, col in enumerate(df.columns, start=1):
            max_len = max(df[col].astype(str).map(len).max(), len(col))
            ws.column_dimensions[get_column_letter(idx)].width = max_len + 2
    output.seek(0)

    filename = f"DevolucionesPorProducto_{start}_a_{end}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# Formulario para ventas por producto
@reportes_bp.route('/ventas_por_producto', methods=['GET'])
@login_required
@rol_requerido('administrador', 'semiadmin')
def ventas_por_producto():
    return render_template('reportes/ventas_por_producto.html')

# Exportación a Excel de ventas por producto
@reportes_bp.route('/ventas_por_producto/export', methods=['GET'])
@login_required
@rol_requerido('administrador', 'semiadmin')
def export_ventas_producto_excel():
    from io import BytesIO
    from openpyxl.utils import get_column_letter
    import pandas as pd

    # Leer fechas
    start = request.args.get('start', '').strip()
    end = request.args.get('end', '').strip()

    try:
        start_date = datetime.strptime(start, '%Y-%m-%d').date()
        end_date = datetime.strptime(end, '%Y-%m-%d').date()
    except ValueError:
        flash('Rango de fechas inválido.', 'warning')
        return redirect(url_for('reportes.ventas_por_producto'))

    # Extraer ventas en el rango
    ventas = BDVenta.query.filter(
        BDVenta.fecha >= start_date,
        BDVenta.fecha <= end_date
    ).all()

    rows = []
    for v in ventas:
        vendedor = Vendedor.query.filter_by(codigo_vendedor=v.codigo_vendedor).first()
        for item in v.items:
            producto = Producto.query.filter_by(codigo=item.producto_cod).first()
            comision_valor = item.comision or Decimal('0')
            pagar_panaderia = item.pagar_pan or Decimal('0')

            rows.append({
                'Fecha': v.fecha,
                'Año': v.fecha.year,
                'Mes': v.fecha.month,
                'Día': v.fecha.day,
                'Día de semana': v.fecha.strftime('%A'),
                'Código producto': item.producto_cod,
                'Nombre producto': producto.nombre if producto else 'Desconocido',
                'Cantidad': item.cantidad,
                'Subtotal': float(item.subtotal),
                'Valor comisión': float(item.comision),
                'Pagar a la Panadería': float(item.pagar_pan),
                'Código vendedor': v.codigo_vendedor,
                'Nombre vendedor': vendedor.nombre if vendedor else 'Desconocido'
            })

    df = pd.DataFrame(rows)
    if df.empty:
        flash('No hay datos en ese rango.', 'info')
        return redirect(url_for('reportes.ventas_por_producto'))

    df['Fecha'] = pd.to_datetime(df['Fecha'])

    # Exportar a Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='VentasPorProducto')
        ws = writer.sheets['VentasPorProducto']
        for idx, col in enumerate(df.columns, start=1):
            max_len = max(df[col].astype(str).map(len).max(), len(col))
            ws.column_dimensions[get_column_letter(idx)].width = max_len + 2
    output.seek(0)

    filename = f"VentasPorProducto_{start}_a_{end}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


